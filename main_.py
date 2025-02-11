# ============================= #
# ======== Bibliotecas ======== #
# ============================= #

import streamlit as st
import zipfile
import tempfile
import os
import re
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import numbers, Border, Side
from openpyxl.utils import get_column_letter
from pandas.api.types import is_float_dtype
from pandas.tseries.offsets import BDay

st.set_page_config(layout="wide")

# ============================= #
# ========== Funções ========== #
# ============================= #

def tratar_valores(valor):
    valor = valor.strip()
    if valor.startswith('(') and valor.endswith(')'):
        valor = valor.replace('(', '').replace(')', '')
        return -float(valor.replace('.', '').replace(',', '.'))
    return float(valor.replace('.', '').replace(',', '.'))

def ler(caminho):
    """
    Lê um arquivo CSV ou XLSX a partir do objeto de arquivo (upload via Streamlit).
    """
    if caminho.name.endswith('.csv'):
        temp1 = pd.read_csv(
            caminho,
            encoding='latin-1',
            low_memory=False,
            decimal=',',
            sep=';'
        )
    elif caminho.name.endswith('.xlsx'):
        temp1 = pd.read_excel(
            caminho,
            engine='openpyxl'
        )
    else:
        raise ValueError("Formato de arquivo não suportado. Envie um arquivo CSV ou XLSX.")
    return temp1

def calcular_atv_probl(row):
    if row['ATRASO'] > 90:
        proporcao = row['PDD_PROPORCIONAL'] / row['VP_PROPORCIONAL']
        if proporcao < 0.2:
            return 1.5
        elif proporcao < 0.5:
            return 1
        else:
            return 0.5
    else:
        return "N"

def classificar_pessoa(doc_sacado):
    if len(doc_sacado) == 18:
        return "PJ"
    else:
        return "PF"

def atualizar_rwa(dataframe):
    dataframe['RWA'] = dataframe['SALDO'] * dataframe['FPR']
    return dataframe

def ativ_probl(row, data):
    data = pd.to_datetime(data, format='%d/%m/%Y', dayfirst=True)
    if (data - pd.to_datetime(row['DATA_EMISSAO_2'])).days > 90:
        return 'S'
    else:
        return 'N'

@st.cache_data
def validar_data(data):
    """
    Valida se a string informada segue o formato dd/mm/YYYY.
    """
    return bool(re.match(r"^\d{2}/\d{2}/\d{4}$", data))

def extrair_zip(zip_file_obj):
    """
    Recebe um objeto BytesIO com os dados do ZIP e extrai seu conteúdo
    para um diretório temporário. Retorna o caminho de extração e o objeto TemporaryDirectory.
    """
    temp_dir = tempfile.TemporaryDirectory()
    temp_path = temp_dir.name
    with zipfile.ZipFile(zip_file_obj) as zip_ref:
        zip_ref.extractall(temp_path)
    return temp_path, temp_dir

def valores_únicos(df, coluna):
    try:
        valores = df[coluna].dropna().unique()
        return valores
    except Exception as e:
        st.error(f"Erro ao extrair valores únicos: {e}")
        return []

# ============================= #
# =========== Código ========== #
# ============================= #

st.title('Cálculo FPR - Extração do arquivo posição')
uploaded_zip = st.file_uploader("Faça o upload do arquivo ZIP contendo os arquivos:", type=["zip"])

if uploaded_zip is not None:
    zip_bytes = BytesIO(uploaded_zip.read())
    
    extraction_path, temp_dir_obj = extrair_zip(zip_bytes)

    extracted_files = os.listdir(extraction_path)
    
    dict = {}
    for arquivo in extracted_files:
        try:
            if re.match(r"^\d+", arquivo) and arquivo.endswith('.csv'):
                key = re.search(r"-(.*)\.csv$", arquivo).group(1)
                caminho_arquivo = os.path.join(extraction_path, arquivo)
                df_temp = pd.read_csv(caminho_arquivo, sep=';', encoding='latin1', on_bad_lines='skip')
                dict[key] = df_temp
        except Exception as e:
            st.error(f"Erro ao processar o arquivo '{arquivo}': {e}")
    
    valores = valores_únicos(df=dict['Renda_Fixa'], coluna='CARTEIRA')
    print('Valores únicos(Fundos): ')
    print(valores)
    print()

    valor_inicial = "FIDC REAG H Y" if "FIDC REAG H Y" in valores else valores[0]
    if valores.size > 0:

        with st.form("form_opcoes"):
            col_fund1, col_fund2, col_fund3, col_fund4 = st.columns(4)
            with col_fund1:
                opção_selecionada = st.selectbox(
                    'Fundo:',
                    valores,
                    index=list(valores).index(valor_inicial) if valor_inicial in valores else 0
                )
            with col_fund2:
                data_mes = st.text_input(
                    'Data:',
                    placeholder="Exemplo: 31/12/2024"
                )
            with col_fund3:
                cota = st.selectbox("Cota:", options=['SUB', 'SR'], index=0)
            with col_fund4:
                investimento = st.number_input("Investimento:",
                                            min_value=0.0,
                                            step=0.01,
                                            format="%.2f")
            col_fund5, col_fund6 = st.columns(2)
            with col_fund5:
                uploaded_csv = st.file_uploader("Estoque:", type=["csv", "xlsx"])
            with col_fund6:
                uploaded_excel = st.file_uploader("Posição:", type=["xlsx", "csv", "xls"])
            submit_button = st.form_submit_button("Executar Análise")

        if submit_button:
            st.session_state["cota"] = cota
            st.session_state["investimento"] = investimento

#### ---- Obtenção do PL ---- ####
        try:
            # RENDA FIXA
            renda_fixa = dict['Renda_Fixa']
            renda_fixa = renda_fixa[renda_fixa["CARTEIRA"] == opção_selecionada]
            
            if renda_fixa['CARTEIRA'].str.contains('FIDC CONSIG PUB').any():
                renda_fixa['TITULO'] = renda_fixa['TITULO'].replace({'COTA MEZANINO': 'COTA SENIOR'})

            renda_fixa = renda_fixa[['TITULO', 'VALORLIQUIDO']]
            tabela_tr = str.maketrans({'.':'', ',':'.', '(':'', ')':''})
            renda_fixa.loc[:, 'VALORLIQUIDO'] = renda_fixa['VALORLIQUIDO'].astype(str).str.translate(tabela_tr).astype(float)
            renda_fixa_grouped = renda_fixa.groupby('TITULO', as_index=False)['VALORLIQUIDO'].sum()

            patrimonio = dict['Patrimonio-Totais']
            patrimonio = patrimonio[patrimonio['CARTEIRA'] == opção_selecionada]  
            patrimonio['VALORPATRIMONIOLIQUIDO'] = patrimonio['VALORPATRIMONIOLIQUIDO']\
        .astype(str)\
        .apply(tratar_valores)
            patrimonio = patrimonio.reset_index(drop=True)
            junior = patrimonio.loc[0, 'VALORPATRIMONIOLIQUIDO']

            titulos_desejados = ['COTA MEZANINO', 'COTA SENIOR']
            renda_fixa_filtrada = renda_fixa_grouped[renda_fixa_grouped['TITULO'].isin(titulos_desejados)]
            soma_total = renda_fixa_filtrada['VALORLIQUIDO'].sum() + junior

            print('Valor do PL Final:', soma_total)
            print()

            df_pl = renda_fixa_filtrada.set_index('TITULO')
            df_pl.loc['COTA JUNIOR'] = [junior]
            df_pl.loc['PL'] = [soma_total]
            print('DataFrame do PL:')
            print(df_pl)
            print()

    #### ---- Obtenção do Saldo ---- ####

            if uploaded_csv is not None:
                df1 = ler(uploaded_csv)
                st.success("Arquivo carregado com sucesso.")
            else:
                st.warning("Aguardando o carregamento dos dados.")
            
            combined_df = df1

            Documento = combined_df['DOC_SACADO']
            columns_to_convert = ['VALOR_NOMINAL', 'VALOR_PRESENTE', 'VALOR_AQUISICAO', 'VALOR_PDD']

            for col in columns_to_convert:
                if not is_float_dtype(combined_df[col]):
                    combined_df[col] = (
                        combined_df[col]
                        .astype(float)
                    )

            soma_valores = combined_df
            soma_valores['DATA_REFERENCIA'] = pd.to_datetime(soma_valores["DATA_REFERENCIA"])
            soma_valores['DATA_VENCIMENTO_AJUSTADA_2'] = pd.to_datetime(soma_valores['DATA_VENCIMENTO_AJUSTADA_2'])
            soma_valores['ATRASO'] = (soma_valores['DATA_REFERENCIA'] - soma_valores['DATA_VENCIMENTO_AJUSTADA_2']).dt.days
            soma_valores['DATA_EMISSAO_2'] = pd.to_datetime(soma_valores['DATA_EMISSAO_2'])
            soma_valores['Ativo problemático'] = soma_valores.apply(ativ_probl, axis=1, data=data_mes)
    # CONTRATO:
            over_90_saldo = soma_valores.loc[soma_valores['ATRASO'] > 90, 'VALOR_PRESENTE'].sum()
            below_90_saldo = soma_valores.loc[soma_valores['ATRASO'] <= 90, 'VALOR_PRESENTE'].sum()
            total_saldo = over_90_saldo + below_90_saldo
            over_90_pdd = soma_valores.loc[soma_valores['ATRASO'] > 90, 'VALOR_PDD'].sum()
            below_90_pdd = soma_valores.loc[soma_valores['ATRASO'] <= 90, 'VALOR_PDD'].sum()
            total_pdd = over_90_pdd + below_90_pdd
            contrato = total_saldo - total_pdd
            print('Contratos: ', contrato.round(2))
            print()

    # CAIXA
            patrimonio['SALDOCAIXAATUAL'] = patrimonio['SALDOCAIXAATUAL']\
        .astype(str)\
        .apply(tratar_valores)
            caixa = patrimonio['SALDOCAIXAATUAL'].iloc[0] 
            print('Caixa: ', caixa)
            print()

    # CONTAS Á RECEBER E Á PAGAR
            try:
                contas = dict['CPR-Lancamentos']
                contas = contas[contas["CARTEIRA"] == opção_selecionada]
                contas = contas[['MODALIDADE', 'VALOR']]
                contas.loc[:, 'VALOR'] = contas.loc[:, 'VALOR'].apply(tratar_valores)
                contas_resultados = contas.groupby(by='MODALIDADE')['VALOR'].sum()
                pagar = contas_resultados.get("Pagar", 0)
                receber = contas_resultados.get("Receber", 0)
            except KeyError:
                pagar = 0
                receber = 0
            
            print("Pagar:", pagar)
            print("Receber:", receber)
            print()

    # FUNDO

            if "Fundos-Fundos" in dict:
                fundos = dict['Fundos-Fundos']
                fundos = fundos[fundos["CARTEIRA"] == opção_selecionada]
                if all(col in fundos.columns for col in ['CODIGO', 'VALORLIQUIDO']):
                    df_fundos = fundos[['CODIGO', 'VALORLIQUIDO']]
                    df_fundos.loc[:, 'VALORLIQUIDO'] = df_fundos.loc[:, 'VALORLIQUIDO'].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False).astype(float)
                    fundo = df_fundos['VALORLIQUIDO'].sum()
                else:
                    print("As colunas 'CODIGO' e 'VALORLIQUIDO' não existem no arquivo 'Fundos-Fundos'.")
                    fundo = 0
            else:
                print("Não houve arquivo 'Fundos-Fundos' no diretório.")
                fundo = 0
            print('Valor Total envolvendo os fundos: ', fundo)

    # RENDA FIXA
            renda_fixa_nao_desejada = renda_fixa_grouped[~renda_fixa_grouped['TITULO'].isin(titulos_desejados)]
            renda = renda_fixa_nao_desejada['VALORLIQUIDO'].sum()
            print('Outras Rendas: ', renda)

    # OUTROS ATIVOS
            if 'Outros_Ativos' in dict:
                ativos = dict['Outros_Ativos']
                ativos = ativos[ativos["CARTEIRA"] == opção_selecionada]
                ativos_não_desejados = ['A VENCER', 'VENCIDOS', 'PDD']
                ativos_desejados = ativos[~ativos['ATIVO'].isin(ativos_não_desejados)]
                ativos_desejados.loc[:, 'VALOR'] = ativos_desejados.loc[:, 'VALOR'].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False).astype(float)
                outros_ativos = ativos_desejados['VALOR'].sum()
            else:
                print('Não há outros ativos á serem adicionados')
                outros_ativos = 0
            print('Soma dos outros ativos: ', outros_ativos)

    # DATAFRAME DF_SALDO
            data = []
            data.append(('Contrato', contrato.round(2)))
            data.append(('Caixa', caixa))
            data.append(('Pagar', pagar))
            data.append(('Receber', receber))
            if 'Fundos-Fundos' in dict:
                for _, row in df_fundos.iterrows():
                    data.append((row['CODIGO'], row['VALORLIQUIDO']))
            else:
                data.append(('Fundo', fundo))
            if not renda_fixa_nao_desejada.empty:
                for _, row in renda_fixa_nao_desejada.iterrows():
                    data.append((row['TITULO'], row['VALORLIQUIDO']))
            else:
                data.append(('Renda Fixa', renda))

            if 'Outros_Ativos' in dict:
                if not ativos_desejados.empty:
                    for _, row in ativos_desejados.iterrows():
                        data.append((row['ATIVO'], row['VALOR']))
                else:
                    data.append(('Outros Ativos', outros_ativos))
            else:
                    data.append(('Outros Ativos', outros_ativos))

            df_saldo = pd.DataFrame(data, columns=['Index', 'SALDO'])
            df_saldo.set_index('Index', inplace=True)
            print("DataFrame Saldo:")
            print(df_saldo)

    # COMPARAÇÃO ENTRE OS VALORES:
            print('# ---- Comparação entre valores ---- #')
            print()
            if df_saldo['SALDO'].sum().round(2) == soma_total.round(2):
                print("VALORES IGUAIS.")

            else:
                print("OS VALORES SÃO DIFERENTES.")
                diferenca = df_saldo['SALDO'].sum().round(2) - soma_total.round(2)
                print(f"A DIFERENÇA É DE: {diferenca:,.2f}")
            
    #### ---- Obtenção do DataBase ---- ####
            investimento = investimento
            valor_investido = investimento
            valor_investido = float(valor_investido)
            valor_investido
            porcentagem = valor_investido / df_saldo['SALDO'].sum()

    # PORCENTAGEM
            porcentagem = porcentagem 
            print('Valor da Porcentagem: ', porcentagem)
            dicionario = {}
            dicionario['Porcentagem'] = porcentagem

    # VALOR PRESENTE PROPORCIONAL
            soma_valores['TIPO_RECEBIVEL'] = soma_valores['TIPO_RECEBIVEL'].apply(lambda x: x.encode('latin1').decode('utf-8'))
            soma_valores['VP_PROPORCIONAL'] = soma_valores['VALOR_PRESENTE'] * porcentagem

    # VALOR PDD PROPORCIONAL
            soma_valores['PDD_PROPORCIONAL'] = soma_valores['VALOR_PDD'] * porcentagem

    # ATIVOS PROBLEMÁTICOS
            soma_valores.loc[:, 'ATV_PROBL'] = soma_valores.apply(calcular_atv_probl, axis=1)

    # PF/PJ
            soma_valores.loc[:, 'PF/PJ'] = soma_valores['DOC_SACADO'].astype(str).apply(classificar_pessoa)

    # FPR
            soma_valores['FPR'] = np.where(
                    soma_valores['TIPO_RECEBIVEL'] == "Precatórios", 12.5,
                    np.where(
                        soma_valores['TIPO_RECEBIVEL'] == "Ação Judicial", 12.5,
                        np.where(
                            soma_valores['ATV_PROBL'] != "N", soma_valores['ATV_PROBL'],
                            np.where(
                                soma_valores['VP_PROPORCIONAL'].astype(float) > 100000,
                                np.where(soma_valores['PF/PJ'] == "PJ", 0.85, 1.00),
                                0.75
                            )
                        )
                    )
                )
            
    # RWA
            soma_valores['RWACPAD']  = (soma_valores['VP_PROPORCIONAL'] - soma_valores['PDD_PROPORCIONAL']) * soma_valores['FPR']

    #### ---- Cálculo do FPR e RWA ---- ####

    # FPR
            dicionário = {
                'Contrato': soma_valores['RWACPAD'].sum() / (soma_valores['VP_PROPORCIONAL'] - soma_valores['PDD_PROPORCIONAL']).sum(),
                'Caixa': 0.2,
                'Pagar': 0.0,
                'Receber': 1.0,
            }
            df_saldo['FPR'] = df_saldo.index.map(dicionário)

    # RWA
            df_saldo['RWA'] = df_saldo['SALDO'] * df_saldo['FPR']

    # DATAFRAME
            print("DataFrame Saldo Atualizado:")
            print(df_saldo)

    #### ---- Obtenção do FPR Final Apurado ---- ####
            def calcular_fpr_final(df_saldo, df_pl, soma_total, over_90_saldo, over_90_pdd, pagar, receber, cota):
                outros = 0
                if 'Ajuste' in df_saldo.index and pd.notna(df_saldo.loc['Ajuste', 'SALDO'] and df_saldo.loc['Ajuste', 'SALDO'] != 0):
                    outros = float(df_saldo.loc['Ajuste', 'SALDO'])
                cota = cota
                sub = None
                sr = None

                if 'CARTEIRA' in patrimonio.columns:  
                    if patrimonio['CARTEIRA'].str.contains('FIDC REAG H Y').any():
                        if cota == 'SUB':
                            sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                            sr = (df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO'] + df_pl.loc['COTA SENIOR', 'VALORLIQUIDO']) / soma_total
                        else:
                            sub = (df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] + df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO']) / soma_total
                            sr = df_pl.loc['COTA SENIOR', 'VALORLIQUIDO'] / soma_total
                    else:
                        if 'COTA MEZANINO' in df_pl.index:
                            if 'COTA SENIOR' in df_pl.index:
                                if cota == 'SUB':
                                    sub = (df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] + df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO']) / soma_total
                                    sr = (df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO'] + df_pl.loc['COTA SENIOR', 'VALORLIQUIDO']) / soma_total
                                else:
                                    sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                                    sr = df_pl.loc['COTA SENIOR', 'VALORLIQUIDO'] / soma_total
                            else:
                                if cota == 'SUB':
                                    # sub = (df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] + df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO']) / soma_total
                                    sub = (df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO']) / soma_total
                                    sr = df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO'] / soma_total
                                else:
                                    sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                                    sr = df_pl.loc['COTA MEZANINO', 'VALORLIQUIDO'] / soma_total
                        else:
                            if 'COTA SENIOR' in df_pl.index:
                                if cota == 'SUB':
                                    sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                                    sr = df_pl.loc['COTA SENIOR', 'VALORLIQUIDO'] / soma_total
                                else:
                                    sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                                    sr = df_pl.loc['COTA SENIOR', 'VALORLIQUIDO'] / soma_total
                            else:
                                if cota == 'SUB':
                                    sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                                    sr = (df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total) - 1
                                else:
                                    sub = df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total
                                    sr = (df_pl.loc['COTA JUNIOR', 'VALORLIQUIDO'] / soma_total) - 1
                else:
                    print("A coluna 'CARTEIRA' não existe no DataFrame.")

                if 'sub' in locals():
                    print('SUB:', sub.round(2))
                    dicionario['SUB'] = sub
                if 'sr' in locals():
                    print('SR:', sr.round(2))
                    dicionario['SR'] = sr

    # Cálculo da Razão da Inadimplência
                razao_inadimplencia = ((over_90_saldo - over_90_pdd) + outros) / df_saldo['SALDO'].sum()
                dicionario['Inadimplência'] = razao_inadimplencia

    # Valores para o Encaixe e Desencaixe
                if cota == 'SR':
                    ponto_encaixe = sub
                    ponto_desencaixe = 1.0
                elif cota == 'SUB':
                    ponto_encaixe = 0.0
                    ponto_desencaixe = sub
                
                dicionario['Encaixe'] = ponto_encaixe
                dicionario['Desencaixe'] = ponto_desencaixe

                print('Ponto de Encaixe: ', ponto_encaixe, 'Ponto de Desencaixe: ', ponto_desencaixe, 'Razão da Inadimplência: ', razao_inadimplencia.round(4))
                print('SUB:', sub, 'SR:', sr)

    # Variáveis para o Cálculo do FPR
                A = ponto_encaixe
                dicionario['A'] =  A
                D = ponto_desencaixe
                dicionario['D'] = D
                W = razao_inadimplencia
                dicionario['W'] = W
                RWAhip = df_saldo['RWA'].sum()
                dicionario['RWAhip'] = RWAhip
                V = (df_saldo['SALDO'].sum() - pagar).round(0)
                dicionario['V'] = V
                F = 0.08

    # Cálculo de Ksa
                Ksa = (RWAhip * F) / V
                dicionario['Ksa'] = Ksa
                print(f"Ksa: {Ksa:.2%}")

    # Cálculo do Ka
                Ka = ((1 - W) * Ksa + (W * 0.5))
                print(f"Ka: {Ka:.2%}")
                dicionario['Ka'] = Ka

                parte_1 = np.exp(((- 1 / Ka) * (D - Ka)))
                print(f"{parte_1:.2%}")
                parte_2 = np.exp((-1 / Ka) * np.maximum(A - Ka, 0))
                print(f"{parte_2:.2%}")
                parte_3 = - (1 / Ka) * (D - Ka - np.maximum(A - Ka, 0))
                print(f"{parte_3:.2%}")
                Kssfa = (parte_1 - parte_2) / parte_3
                print(f"{Kssfa:.2%}")
                dicionario['Kssfa'] = Kssfa
                parte_4 = ((Ka - A) / (D - A)) * (1/F)
                print(f"{parte_4:.2%}")
                parte_5 = ((D - Ka) / (D - A)) * (1/F) * Kssfa
                print(f"{parte_5:.2%}")
                FPR = parte_4 + parte_5
                print(f"{FPR:.2%}")
                dicionario['FPR'] = FPR

                Ka = Ka
                I = (1/F)
                dicionario['I'] = I
                print(f"{I:.2%}")
                II = (1/F) * Kssfa
                dicionario['II'] = II
                print(f"{II:.2%}")
                III = FPR
                dicionario['III'] = III
                print(f"{III:.2%}")

                A = ponto_encaixe
                D = ponto_desencaixe 
                W = razao_inadimplencia
                RWAhip = df_saldo['RWA'].sum()
                V = (df_saldo['SALDO'].sum() - pagar).round(0)

                print(f'Verifica se {D} é menor que {float(Ka)}')
                if D <= float(Ka):
                    D6 = True
                    print(f'Caso D6: o valor do D é igual á {D} e o valor do Ka é {Ka}, nesse caso {D} é menor que {float(Ka)}, D6 passar á ser {D6}')
                else:
                    D6 = False
                    print(f'Caso do {D} ser maior que {float(Ka)}, D6 é, portanto, {D6}')

                print()
                print(f'Verifica se {float(A)} é maior que {float(Ka)}')
                if float(A) >= float(Ka):
                    D7 = True
                    print(f'Caso D7: o valor do A é igual á {float(A)} e o valor do Ka é {Ka}, nesse caso {float(A)} é maior que {float(Ka)}, D7 passa a ser {D7}')
                else:
                    D7 = False
                    print(f'Caso do {float(Ka)} ser menor que {float(Ka)}, D7 é, portanto, {D7}')

                F7 = 0.25

                print()

                if D6 == True:
                    F6 = float(I)
                    print(f'Se {D6} == True, então F6 é igual á {float(I)}')
                else:
                    if D7 == True:
                        F6 = float(II)
                        print(f'Se {D7} == True, então F6 é igual á {float(II)}')
                    else:
                        F6 = float(III)
                        print(f'Se {D7} == False, então F6 é igual á {float(III)}')

                print(F6, F7)
                print('Pegar o valor máximo:')
                print()
                FPR_apurado = float(np.maximum(F6, F7))
                print()
                print()
                print('# ---- Valor Final para o FPR ---- #')
                print(f"{FPR_apurado:.0%}")
                return FPR_apurado

    #### ---- Interface do Streamlit ---- #####
            st.title("Edição do DataFrame")

            columns_to_display = ['Index', 'SALDO', 'FPR']
            editable_columns = df_saldo.reset_index()[columns_to_display]

            num_blank_rows = 10
            blank_rows = pd.DataFrame([{'Index': '', 'SALDO': 0, 'FPR': None}] * num_blank_rows)
            editable_columns = pd.concat([editable_columns, blank_rows], ignore_index=True)

            st.markdown(
            """
            <hr style="border: 1px solid black; margin: 20px 0;">
            """, unsafe_allow_html=True)

            st.write("## 1) Edição dos Valores da Coluna FPR:")
            edited_subset = st.data_editor(editable_columns, use_container_width=True)

            updated_df = edited_subset[edited_subset['Index'] != ''].set_index('Index')
            updated_df['SALDO'] = updated_df['SALDO'].astype(float)
            updated_df['FPR'] = updated_df['FPR'].astype(float)

    # Recalcular a coluna RWA
            updated_df['RWA'] = updated_df['SALDO'] * updated_df['FPR']

    # Atualizar o DataFrame original
            df_saldo = updated_df
            df_saldo = df_saldo[~df_saldo.index.isin(["Renda Fixa", "Outros Ativos"])]
            
            st.markdown(
            """
            <hr style="border: 1px solid black; margin: 20px 0;">
            """, unsafe_allow_html=True)
            st.write("## 2) Comparação entre DataFrames:")
            fpr_final = calcular_fpr_final(
                df_saldo, df_pl, soma_total, over_90_saldo, over_90_pdd, pagar, receber, cota
            )
            col1, col2 = st.columns([1, 2])

            with col1:
                if not df_pl.empty:
                    st.write("#### DataFrame Patrimônio Líquido")
                    st.dataframe(df_pl)
                    st.write("#### DataFrame Saldo")
                    st.dataframe(df_saldo)
                    # st.write(f'Total da coluna Saldo: {locale.format_string('%.2f', df_saldo['SALDO'].sum(), grouping=True)}')
                # st.write(f'Toda da coluna RWA: {locale.format_string('%.2f', df_saldo['RWA'].sum(), grouping=True)}')
                expo = df_saldo['SALDO'].sum() - pagar
                # st.write(f'EXPO: {locale.format_string('%.2f', expo, grouping=True)}')
                st.markdown(
                f"""
                <div style="text-align: center; font-size: 24px; font-weight: bold; color: black;">
                Comparação entre os valores:
                </div>
                """,
                unsafe_allow_html=True,)
                if df_saldo['SALDO'].sum().round(2) == soma_total.round(2):
                    print("VALORES IGUAIS.")
                    st.markdown(
                    """
                    <div style="text-align: center; font-size: 16px; font-weight: bold; color: green;">
                    OS VALORES ENTRE OS DATAFRAMES SÃO IGUAIS.
                    </div>
                    """,
                    unsafe_allow_html=True,)
                else:
                    print("OS VALORES SÃO DIFERENTES.")
                    diferenca = df_saldo['SALDO'].sum().round(2) - soma_total.round(2)
                    print(f"A DIFERENÇA É DE: {diferenca:,.2f}")
                    st.markdown(
                    f"""
                    <div style="text-align: center; font-size: 16px; font-weight: bold; color: red;">
                    OS VALORES ENTRE OS DATAFRAMES SÃO DIFERENTES, DIFERENÇA DE: {diferenca:.2f}
                    </div>
                    """,
                    unsafe_allow_html=True,)

            with col2:
                st.write('### Database:')
                st.dataframe(soma_valores)

            st.write('#### Saldo %:')
            saldo_porcentagem = df_saldo['SALDO'] * porcentagem
            st.dataframe(saldo_porcentagem.to_frame().T)
            st.write('### Informações:')
            df_dicionário = pd.DataFrame(dicionario.items(), columns=['Item', 'Valor dos Itens'])
            df_dicionário.set_index('Item', inplace=True)
            df_dicionário['Valor dos Itens'] = df_dicionário['Valor dos Itens'].round(3)
            st.dataframe(df_dicionário.T)
            st.markdown(
            """
            <hr style="border: 1px solid black; margin: 20px 0;">
            """,
            unsafe_allow_html=True)
            st.markdown(
                    f"""
                    <div style="text-align: center; font-size: 30px; font-weight: bold;">
                    Valor Final do FPR Apurado: {fpr_final:.2%}
                    </div>
                    """,
                    unsafe_allow_html=True,)
            st.markdown(
            """
            <hr style="border: 1px solid black; margin: 20px 0;">
            """,
            unsafe_allow_html=True)

    #### ---- Arquivo de Posição ---- #####

            st.write("## 3) Arquivo Posição:")
            if uploaded_excel is not None:
                df_excel = pd.read_excel(uploaded_excel)
            else:
                st.warning("Nenhum arquivo Excel foi carregado.")
        
            st.write('### DataFrame Editável:')
            segunda_coluna = df_excel.iloc[:, 1]

            posicao = pd.DataFrame()
            posicao['X'] = soma_valores['VP_PROPORCIONAL']
            posicao['Sistema Origem'] = opção_selecionada
            posicao['Localização'] = 'BR'
            posicao['Tipo de carteira'] = 0
            posicao['Produto'] = 'OPERACAO_CREDITO_FUNDO'
            posicao['Data de contratação'] = soma_valores['DATA_EMISSAO_2'].dt.strftime('%d/%m/%Y').astype(str)
            posicao['Data Vencimento'] = soma_valores['DATA_VENCIMENTO_AJUSTADA_2'].dt.strftime('%d/%m/%Y').astype(str)
            posicao['Moeda Operação'] = 'BRL'
            posicao['Forma de liquidacao'] = np.nan
            posicao['Indexador Ativo'] = np.nan
            posicao['Indexador Passivo'] = np.nan
            posicao['Valor atual'] = soma_valores['VP_PROPORCIONAL']
            posicao['Valor original'] = np.nan
            posicao['Sistema Registro'] = np.nan
            posicao['Tipo Controle'] = np.nan
            posicao['Tipo Pessoa'] = np.nan
            posicao['Grupo Econômico/Matriz/Conectada'] = np.nan
            posicao['Faturamentos'] = np.nan
            posicao['Provisoes'] = pd.to_numeric(soma_valores['PDD_PROPORCIONAL'])
            posicao['Notional'] = np.nan
            posicao['Contraparte'] = df_excel['Unnamed: 20'].mode()[0]
            posicao['Contraparte'] = posicao['Contraparte'].fillna('').astype(str).str.replace(',', '', regex=False) ####
            posicao['Compromissada Over ou operação marcada à mercado'] = np.nan
            posicao['Tipo acordo compensação (CEM)/ Indicador Margem (SA-CCR)'] = np.nan
            posicao['Ativo objeto'] = soma_valores['TIPO_RECEBIVEL']
            posicao['Emitente'] = soma_valores['DOC_SACADO']
            posicao['Emitente'] = posicao['Emitente'].str.replace(r'[./-]', '', regex=True)
            posicao['Indicador MTM Ativo'] = np.nan
            posicao['Data Vencimento Ativo'] = soma_valores['DATA_VENCIMENTO_AJUSTADA_2'].dt.strftime('%d/%m/%Y').astype(str)
            posicao['Valor Ativo Associado'] = soma_valores['VP_PROPORCIONAL']
            posicao['Moeda Ativo Objeto'] = np.nan
            posicao['Nome da Contraparte'] = posicao['Sistema Origem']
            posicao['Nome do Emissor'] = np.nan
            posicao['COSIF'] = np.nan
            posicao['Descrição  ativo objeto'] = np.nan
            posicao['Razão para conectar contraparte'] = np.nan
            posicao['P'] = np.nan
            posicao['Data do próximo fluxo'] = np.nan
            posicao['K'] = np.nan
            posicao['Lambda'] = np.nan
            posicao['Data início ou exercício do ativo'] = np.nan
            posicao['Data da Posição'] = np.nan
            posicao['Ativo problemático'] = soma_valores['Ativo problemático']
            posicao['Característica especial'] = np.nan
            posicao['Campo reservado'] = np.nan
            posicao['Valor atual cenário 1'] = np.nan
            posicao['Valor notional cenário 1'] = np.nan
            posicao['Valor atual cenário 2'] = np.nan
            posicao['Valor notional cenário 2'] = np.nan
            posicao['Valor atual cenário 3'] = np.nan
            posicao['Valor notional cenário 3'] = np.nan
            posicao['Valor atual cenário 4'] = np.nan
            posicao['Valor notional cenário 4'] = np.nan
            posicao['Valor atual cenário 5'] = np.nan
            posicao['Valor notional cenário 5'] = np.nan
            posicao['Valor atual cenário 6'] = np.nan
            posicao['Valor notional cenário 6'] = np.nan
            posicao = posicao.reset_index()
            posicao = posicao.rename({'index': 'Num operacao'}, axis=1)
            posicao['Num operacao'] = posicao['Num operacao'].apply(lambda x: x + 1) 
            posicao = posicao.drop('X', axis=1)

            moda_colunas = posicao.mode().iloc[0]

            for index, valor in df_saldo.iterrows():
                if index == "Contrato":
                    continue
                nova_linha = {
                    "Produto": index,                                
                    "Valor atual": valor['SALDO'] * porcentagem,    
                }

                for coluna in posicao.columns:
                    if coluna not in nova_linha: 
                        nova_linha[coluna] = moda_colunas[coluna]
                        
                nova_linha_df = pd.DataFrame([nova_linha], columns=posicao.columns)
                
                posicao = pd.concat([posicao, nova_linha_df], ignore_index=True)      
            
            ultimo_valor = posicao['Num operacao'].max()
            contador = ultimo_valor # + 1
            indices = posicao.index[posicao['Num operacao'] == 1].tolist()
            for i in indices:
                posicao.at[i, 'Num operacao'] = contador
                contador += 1
            posicao.loc[0, 'Num operacao'] = 1

            posicao['Produto'] = posicao['Produto'].replace("Caixa", "CAIXA_FIDC")
            posicao['Produto'] = posicao['Produto'].replace("Pagar", "VALOR_A_PAGAR_FIDC")
            posicao['Produto'] = posicao['Produto'].replace("Receber", "VALOR_A_RECEBER_FIDC")

            for i in posicao['Produto'].unique():
                if i != "OPERACAO_CREDITO_FUNDO":
                    posicao.loc[posicao['Produto'] == i, 'Data de contratação'] = data_mes
                    proximo_dia_util = (pd.to_datetime(data_mes) + BDay(1)).strftime('%d/%m/%Y')
                    posicao.loc[posicao['Produto'] == i, 'Data Vencimento'] = proximo_dia_util
                    posicao.loc[posicao['Produto'] == i, 'Ativo problemático'] = 'N'
            
            mask = posicao['Produto'].str.startswith('F')
            posicao.loc[mask, 'Sistema Origem'] = posicao.loc[mask, 'Produto']
            posicao.loc[mask, 'Produto'] = "TVM - FUNDO"
            mask_other = (posicao['Produto'] == 'DC FURTADO COELHO AD') | (posicao['Produto'] == 'RECURAÇÃO JUDICIAL')
            posicao.loc[mask_other, 'Produto'] = 'VALOR_A_RECEBER_FIDC'

            edited_posicao = st.data_editor(posicao, num_rows="dynamic", use_container_width=True, height=600)

            st.write("### Arquivo Final:")
            st.write('* Faço o download do arquivo posição.')
            lista = [
                "Data do Ciclo",
                data_mes,
                "Tipo de Movimento",
                "I",
                "Empresa",
                df_excel.columns[5],
                "Alias",
                df_excel.columns[7],
                "Unnamed: 8","Unnamed: 9","Unnamed: 10",
                posicao['Valor atual'].sum(),
                posicao['Valor original'].sum(),
                "Unnamed: 13","Unnamed: 14","Unnamed: 15","Unnamed: 16","Unnamed: 17",
                posicao['Provisoes'].sum(),
                "Unnamed: 19","Unnamed: 20","Unnamed: 21","Unnamed: 22","Unnamed: 23","Unnamed: 24","Unnamed: 25","Unnamed: 26","Unnamed: 27","Unnamed: 28","Unnamed: 29",
                "Unnamed: 30","Unnamed: 31","Unnamed: 32","Unnamed: 33","Unnamed: 34","Unnamed: 35","Unnamed: 36","Unnamed: 37","Unnamed: 38","Unnamed: 39","Unnamed: 40",
                "Unnamed: 41","Unnamed: 42","Unnamed: 43","Unnamed: 44","Unnamed: 45","Unnamed: 46","Unnamed: 47","Unnamed: 48","Unnamed: 49","Unnamed: 50","Unnamed: 51",
                "Unnamed: 52","Unnamed: 53", "Unnamed: 54", "Unnamed: 55"
            ]

            num_colunas = len(edited_posicao.columns)
            if len(lista) < num_colunas:
                lista.extend([""] * (num_colunas - len(lista)))
            elif len(lista) > num_colunas:
                lista = lista[:num_colunas]

            edited_posicao.loc[-1] = edited_posicao.columns 
            edited_posicao.index = edited_posicao.index + 1 
            edited_posicao = edited_posicao.sort_index()  

            edited_posicao.columns = lista

            st.dataframe(edited_posicao)

    #### ---- Edição do Arquivo Posição ---- #####

            uploaded_csv = st.file_uploader("### Faça o upload do arquivo POSIÇÃO", type=["csv"])

            if uploaded_csv is not None:
                df = pd.read_csv(uploaded_csv)
                df.columns = ["" if "Unnamed" in col else col for col in df.columns]

                output = BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name="Sheet1")
                excel_data = output.getvalue()

                temp_excel_path = "temp_arquivo_convertido.xlsx"
                with open(temp_excel_path, "wb") as temp_file:
                    temp_file.write(excel_data)

                wb = load_workbook(temp_excel_path)
                ws = wb.active

                ws.delete_cols(1)

                for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

                for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = numbers.FORMAT_DATE_DDMMYY

                
                coluna_L = 12  
                soma_coluna = 0

                for row in ws.iter_rows(min_row=2, min_col=coluna_L, max_col=coluna_L): 
                    for cell in row:
                        if cell.value is not None:  
                            try:
                                cell.value = float(cell.value) 
                                soma_coluna += cell.value  
                                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  
                            except ValueError:
                                pass 

                ws.cell(row=1, column=coluna_L).value = soma_coluna
                ws.cell(row=1, column=coluna_L).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                
                coluna_D = "D"  
                coluna_D_index = ws[coluna_D + "1"].column  

                for row in ws.iter_rows(min_row=2, min_col=coluna_D_index, max_col=coluna_D_index):
                    for cell in row:
                        if cell.value is not None: 
                            try:
                                valor = float(cell.value)
                                if valor == 0.0:  
                                    cell.value = 0
                                else:
                                    cell.value = valor  
                                cell.number_format = numbers.FORMAT_NUMBER  
                            except ValueError:
                                pass

                coluna_S = "S"
                coluna_S_index = ws[coluna_S + "1"].column  

                soma_coluna_S = 0
                for row in ws.iter_rows(min_row=2, min_col=coluna_S_index, max_col=coluna_S_index):
                    for cell in row:
                        if cell.value not in (None, "0", "0.0", 0.0):  
                            try:
                                cell.value = round(float(cell.value), 2)
                                soma_coluna_S += cell.value  
                                cell.number_format = "#,##0.00"  
                            except ValueError:
                                pass 

                ws[coluna_S + "1"].value = round(soma_coluna_S, 2)
                ws[coluna_S + "1"].number_format = "#,##0.00"

                coluna_AB = "AB"
                coluna_AB_index = ws[coluna_AB + "1"].column  

                for row in ws.iter_rows(min_row=2, min_col=coluna_AB_index, max_col=coluna_AB_index):
                    for cell in row:
                        if cell.value not in (None, "0", "0.0", 0.0):  
                            try:
                                cell.value = round(float(cell.value), 2)
                                cell.number_format = "#,##0.00" 
                            except ValueError:
                                pass 
                
                for cell in ws['AA'][1:]:
                    if cell.value:
                        cell.number_format = 'DD/MM/YYYY'

                ws['M1'] = '=L1-S1'

                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2  
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                borda_vazia = Border(left=Side(border_style=None),
                        right=Side(border_style=None),
                        top=Side(border_style=None),
                        bottom=Side(border_style=None))

                for cell in ws[1]:
                    if cell.value is None:  
                        cell.border = borda_vazia  

                wb.save(temp_excel_path)
                wb.close()

                with open(temp_excel_path, "rb") as modified_file:
                    modified_excel_data = modified_file.read()

                st.download_button(
                    label="Baixar em Excel",
                    data=modified_excel_data,
                    file_name=f"arquivo_convertido_{str(opção_selecionada)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except:
            pass
